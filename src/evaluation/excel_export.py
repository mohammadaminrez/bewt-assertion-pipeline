from __future__ import annotations

"""Export experiment results to Excel for manual annotation and import back."""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..models import ExperimentResult

CLASSIFICATION_OPTIONS = "correct,over_assertive,under_assertive,wrong_assertion,not_executable"

GUIDE_ROWS = [
    ["Classification", "When to use", "Example"],
    ["correct", "Generated checks the same thing as gold (same logic, maybe different syntax)", 'Gold: assertEquals("updater", page.getAccessLevel(2))\nGenerated: assertTrue(page.getAccessLevel(2).equals("updater"))'],
    ["over_assertive", "Generated checks MORE than the gold (extra assertions or stricter checks)", 'Gold: assertEquals("updater", page.getAccessLevel(2))\nGenerated: same + assertEquals("user001", page.getUsername(2))'],
    ["under_assertive", "Generated checks LESS than the gold (weaker, would pass when it should fail)", 'Gold: assertEquals("updater", page.getAccessLevel(2))\nGenerated: assertTrue(driver.getPageSource().contains("updater"))'],
    ["wrong_assertion", "Generated checks the WRONG thing (different element, value, or logic)", 'Gold: assertEquals("updater", page.getAccessLevel(2))\nGenerated: assertEquals("admin", page.getAccessLevel(1))'],
    ["not_executable", "Does not compile or cannot run", "Syntax error, missing method, wrong locator"],
]


def export_results_to_excel(
    results: list[ExperimentResult],
    output_path: Path,
    pre_classifications: dict[str, str] | None = None,
) -> None:
    """Export results to an Excel file for manual annotation.

    Args:
        results: Experiment results to export.
        output_path: Path to write the .xlsx file.
        pre_classifications: Optional dict of "app|class|treatment|model" -> classification
            from LLM pre-classification.
    """
    rows = []
    for r in results:
        key = f"{r.test_record.app}|{r.test_record.class_name}|{r.treatment}|{r.model}"
        pre_class = (pre_classifications or {}).get(key, "")
        rows.append({
            "App": r.test_record.app,
            "Test Class": r.test_record.class_name,
            "Method": r.test_record.method_name,
            "Treatment": r.treatment,
            "Model": r.model,
            "Gold Standard Assertion": r.test_record.gold_standard,
            "Generated Assertion": r.generated_assertion,
            "Compiles": r.compiles,
            "Pass/Fail": "PASS" if r.passes else "FAIL",
            "Fail Reason": r.notes if not r.passes and r.notes and r.notes != "execution skipped" else "",
            "Semantic Sim (heuristic)": round(r.semantic_similarity, 4),
            "Auto-Classification (suggestion)": r.error_category.value,
            "Manual Classification": pre_class,
            "Manual Notes": "",
        })

    df = pd.DataFrame(rows)
    df.sort_values(by=["App", "Test Class", "Treatment", "Model"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(str(output_path), index=False, sheet_name="Results", engine="openpyxl")

    wb = load_workbook(str(output_path))
    ws = wb["Results"]

    _apply_formatting(ws, len(df))
    _add_guide_sheet(wb)

    wb.save(str(output_path))


def _apply_formatting(ws, num_rows: int) -> None:
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    pass_fail_col = None
    manual_class_col = None
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "Pass/Fail":
            pass_fail_col = idx
        if cell.value == "Manual Classification":
            manual_class_col = idx

    for row in range(2, num_rows + 2):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        if pass_fail_col:
            cell = ws.cell(row=row, column=pass_fail_col)
            if cell.value == "PASS":
                cell.fill = pass_fill
            elif cell.value == "FAIL":
                cell.fill = fail_fill

    if manual_class_col:
        dv = DataValidation(
            type="list",
            formula1=f'"{CLASSIFICATION_OPTIONS}"',
            allow_blank=True,
        )
        dv.error = "Please select a valid classification"
        dv.errorTitle = "Invalid classification"
        dv.prompt = "Select classification"
        dv.promptTitle = "Manual Classification"
        ws.add_data_validation(dv)
        for row in range(2, num_rows + 2):
            dv.add(ws.cell(row=row, column=manual_class_col))

    col_widths = {
        "App": 12, "Test Class": 25, "Method": 20,
        "Treatment": 10, "Model": 15,
        "Gold Standard Assertion": 50, "Generated Assertion": 50,
        "Compiles": 10, "Pass/Fail": 10, "Fail Reason": 40,
        "Semantic Sim (heuristic)": 15, "Auto-Classification (suggestion)": 22,
        "Manual Classification": 22, "Manual Notes": 30,
    }
    for idx, cell in enumerate(ws[1], 1):
        width = col_widths.get(cell.value, 15)
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _add_guide_sheet(wb) -> None:
    gs = wb.create_sheet("Classification Guide")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for row_idx, row_data in enumerate(GUIDE_ROWS, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = gs.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font

    gs.column_dimensions["A"].width = 20
    gs.column_dimensions["B"].width = 55
    gs.column_dimensions["C"].width = 65


def import_classifications_from_excel(
    excel_path: Path,
) -> list[dict]:
    """Read manual classifications from an annotated Excel file.

    Returns a list of dicts with keys: app, class_name, treatment, model, manual_classification, manual_notes.
    """
    df = pd.read_excel(str(excel_path), sheet_name="Results", engine="openpyxl")

    annotations = []
    for _, row in df.iterrows():
        manual_class = row.get("Manual Classification", "")
        if pd.isna(manual_class) or not str(manual_class).strip():
            continue
        annotations.append({
            "app": row["App"],
            "class_name": row["Test Class"],
            "treatment": row["Treatment"],
            "model": row["Model"],
            "manual_classification": str(manual_class).strip(),
            "manual_notes": str(row.get("Manual Notes", "") or ""),
        })

    return annotations
